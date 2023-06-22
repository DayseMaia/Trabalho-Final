# Importando módulos e bibliotecas necessárias para funcionamento do código
from flask import *
from bancoDeDados.mainDB import *
import openpyxl

# Criando instância do Flask
app = Flask(__name__)

# Carregando o arquivo "dados.xlsx" usando a biblioteca openpyxl
workbook = openpyxl.load_workbook('dados.xlsx')
planilha = workbook['dados_calculo']

# Rota principal para a página inicial
@app.route('/')
def home():
    return render_template('index.html')

# Rota para a página de simulação
@app.route('/simulacao')
def simulacao():
    return render_template('simulacao.html')

# Rota para a página de simulaçao para sistema On Grid
@app.route('/ongrid', methods=("GET", "POST"))
def ongrid():

    if request.method == "GET":
        return render_template('ongrid.html')

    if request.method == "POST":
        resultado = conexao.manipularBanco(f'''
            INSERT INTO "sistema"
            values(DEFAULT , DEFAULT , 'ONGRID', '{request.form['KWHRDIA']}', '{request.form['ENERGIASOL']}', '{request.form['ESTADO']}', '{request.form['CIDADE']}', 'NÃO UTILIZA BATERIAVOLTS', 'NÃO UTILIZA ENERGIADIARESERVA', '{request.form['PAINELTIPO']}', '{request.form['MODULOWATT']}', '{request.form['TEMPERATURAOPCAO']}', '{request.form['PERDASINVERSOR']}', '{request.form['FATORSEGURANCAINVERSOR']}', '{request.form['PERDASCABO']}', '{request.form['PERDASINCOMPATIBILIDADE']}', '{request.form['PERDASSUJEIRA']}', 'NÃO UTILIZA PROFUNDIDADE', 'NÃO UTILIZA BATERIAEFICIENCIA' )
            ''')
        
        if resultado:
            return redirect(url_for("login"))
        else:
            return "Erro na inserção"

# Rota para a página de simulação para sistema Off Grid
@app.route('/offgrid', methods=("GET", "POST"))
def offgrid():
    if request.method == "GET":
        return render_template('offgrid.html')

    if request.method == "POST":
        resultado = conexao.manipularBanco(f'''
            INSERT INTO "sistema"
            values(DEFAULT , DEFAULT , 'OFFGRID', '{request.form['KWHRDIA']}', 'NÃO UTILIZA ENERGIASOL', '{request.form['ESTADO']}', '{request.form['CIDADE']}', '{request.form['BATERIAVOLTS']}', '{request.form['ENERGIADIARESERVA']}', '{request.form['PAINELTIPO']}', '{request.form['MODULOWATT']}', '{request.form['TEMPERATURAOPCAO']}', '{request.form['PERDASINVERSOR']}', '{request.form['FATORSEGURANCAINVERSOR']}', '{request.form['PERDASCABO']}', '{request.form['PERDASINCOMPATIBILIDADE']}', '{request.form['PERDASSUJEIRA']}', '{request.form['PROFUNDIDADE']}', '{request.form['BATERIAEFICIENCIA']}')
            ''')
        
        if resultado:
            return redirect(url_for("login"))
        else:
            return "Erro na inserção"

# Rota para a página de simulação para sistema Híbrido
@app.route('/hybrid', methods=("GET", "POST"))
def hybrid():
    
    if request.method == "GET":
        return render_template('hybrid.html')

    if request.method == "POST":
        cidade = request.form['CIDADE']
        estado = request.form['ESTADO']
        energiaUsuario = float(request.form['KWHRDIA'])*1000  # energia em Wh
        Eficiencia = 0.8
        potenciaPlaca = 550 # Modiicar para valores do banco
        quantPlaca = 0 
        irradiacaoAnual = 0

        for row1, row2 in zip(planilha.iter_rows(min_row=1, max_row=planilha.max_row, min_col=1, max_col=8),
                      planilha.iter_rows(min_row=1, max_row=planilha.max_row, min_col=1, max_col=8)):
            if row1[3].value == cidade and row2[5].value == estado.upper():
                irradiacaoAnual = float(planilha.cell(row=row1[0].row, column=7).value)
                # print(row1[3].value, cidade, row2[5].value, estado.upper(), irradiacaoAnual)
                break
    
        consumo = energiaUsuario / (Eficiencia * 30 * irradiacaoAnual)
        consumo = consumo * 1000
        quantPlaca = consumo / potenciaPlaca 
        precoPlaca = 1100 # preço sem conisderar instalação e nem bateria e inversor 
        precoProjeto = quantPlaca * precoPlaca
        
        # print(f"ver numero: {quantPlaca}")
        #'{request.form['PAINELTIPO']}'
        #'{request.form['MODULOWATT']}'

        resultado = conexao.manipularBanco(f'''
        INSERT INTO "sistema"
        values(DEFAULT , DEFAULT , 'HYBRID', '{request.form['KWHRDIA']}', '{request.form['ENERGIASOL']}', '{request.form['ESTADO']}', '{request.form['CIDADE']}', '{request.form['BATERIAVOLTS']}', '{request.form['ENERGIADIARESERVA']}', default, default, '{request.form['TEMPERATURAOPCAO']}', '{request.form['PERDASINVERSOR']}', '{request.form['FATORSEGURANCAINVERSOR']}', '{request.form['PERDASCABO']}', '{request.form['PERDASINCOMPATIBILIDADE']}', '{request.form['PERDASSUJEIRA']}', '{request.form['PROFUNDIDADE']}', '{request.form['BATERIAEFICIENCIA']}')
        ''')
    
        # variáveis para rodar na página
        #tamanhoUsina = 15 > 6
        potenciaTotalsistema = "Placa Solar Fotovoltaica 550W Luxen - LNVU-550M" 
        PotenciaUsina = potenciaPlaca * quantPlaca
        PotenciaPlaca = potenciaPlaca
        areaTotal = "2279 x 1134 x 35 mm"
        classificaModulo = "A Placa Solar Series 5 de 550W da marca Luxen Solar \n Possui 144 células de silício monocristalino e tecnologias de alto nível como Half Cut e \n A dopagem de gálio para maior eficiência no longo prazo. \n Tem garantia total de 12 anos, \n Já incluindo os 90 dias legais contra defeito de fabricação."
        numModulos = f"{quantPlaca:.0f}"
        #capacidadeBateria = "resposta14"
        potenciaInversor = "resposta15"
        CustoTotal = f"R$ {precoPlaca:.2f}"
        custoSubGov = f"R$ {precoProjeto:.2f}"
        if resultado:
            return render_template('hybrid.html',potenciaTotalsistema = potenciaTotalsistema, PotenciaPlaca= PotenciaPlaca,PotenciaUsina=PotenciaUsina,areaTotal=areaTotal,classificaModulo=classificaModulo,numModulos=numModulos,capacidadeBateria=500,potenciaInversor=potenciaInversor,CustoTotal=CustoTotal,custoSubGov=custoSubGov) # MODIFICAÇÃO 
            # return redirect(url_for("login"))
        else:
            return "Erro na inserção"

# Rota para a página de login
@app.route('/login', methods=['GET', 'POST'])
def login():

    if request.method == "GET":
        return render_template('login.html')
    
    if request.method == "POST":
        # nome = request.form["nome"]
        # sobrenome = request.form['sobrenome']
        email = request.form['email']
        senha = request.form["senha"]
        
        # conn = connect_to_db()
        # cur = conn.cursor()
        # cur.execute("SELECT * FROM users WHERE nome = %s AND sobrenome = %s AND email = %s AND senha = %s", (nome, sobrenome, email, senha))
        # //////////fazer um select para conferir os dados.
        resultado = conexao.consultarBanco(f'''SELECT "registro_email","registro_senha" FROM "registro" WHERE "registro_email" = '{email}' ;''')
        user = False
        if resultado:
            resultado = resultado[0] # tirar da primeira lista os resultados
            emailDB = resultado[0]
            senhaDB = resultado[1]
            print("email",emailDB,"senha",senhaDB)
            print(email,senha)
            if email == emailDB and senha == senhaDB:
                user = True

        # user = cur.fetchone()
        # cur.close()
        # conn.close()

        if user:
            # session["nome"] = user[1]  # Armazena o nome de usuário na sessão
            return redirect("/address")
        else:
            return render_template('login.html', error="Credenciais inválidas. Tente novamente.")

    return render_template('login.html')

# Rota para a página de registro de um novo usuário
@app.route("/register", methods=["GET", "POST"])
def register():

    if request.method == "POST":
        nome = request.form["nome"]
        sobrenome = request.form['sobrenome']
        email = request.form['email']
        senha = request.form["senha"]

        # conn = connect_to_db()
        # cur = conn.cursor()
        # cur.execute("INSERT INTO users (nome, sobrenome, email, senha) VALUES (%s, %s, %s, %s)", (nome, sobrenome, email, senha))
        resultado = conexao.manipularBanco(f'''
            INSERT INTO "registro"
            values(DEFAULT , DEFAULT , '{nome}', '{sobrenome}', '{email}', '{senha}')
            ''')
        # conn.commit()
        # cur.close()
        # conn.close()

        return redirect('/login')
    
    return render_template("register.html")

# Rota para a página de cadastro de endereço
@app.route('/address', methods=("GET", "POST"))
def address():

    if request.method == "GET":
        return render_template('address.html')
    
    if request.method == "POST":
        resultado = conexao.manipularBanco(f'''
            INSERT INTO "endereco"
            values( DEFAULT , DEFAULT ,'{request.form['CEP']}','{request.form['RUA']}', '{request.form['NUMERO']}', '{request.form['COMPLEMENTO']}', '{request.form['BAIRRO']}', '{request.form['CIDADE']}', '{request.form['ESTADO']}')
            ''')
        
        if resultado:
            return redirect(url_for("pay"))
        else:
            return "Erro na inserção"

# Rota para fazer logout
@app.route("/logout")
def logout():
    session.pop("nome", None)
    return redirect("/login")

# Rota para a página de pagamento
@app.route('/pay', methods=("GET", "POST"))
def pay():

    if request.method == "GET":
        return render_template('pay.html')

    if request.method == "POST":
        resultado = conexao.manipularBanco(f'''
            INSERT INTO "cartao"
            values( DEFAULT , DEFAULT ,'{request.form['NUMERO']}','{request.form['TITULAR']}', '{request.form['MES']}', '{request.form['ANO']}', '{request.form['CVV']}')
            ''')
        
        if resultado:
            return redirect(url_for("home"))
        else:
            return "Erro na inserção"

# Executa a aplicação Flask
if __name__ == "__main__":
    app.run(debug=True)
